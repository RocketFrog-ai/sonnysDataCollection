"""
Extract structured data from the Sonny's proforma Excel workbooks in
experiments/old-proforma-analysis/old-proforma-data/ into a single combined CSV.

Two Excel layouts exist:
  * "Input Form" layout  (112 xlsx) -- clean template, sheet named "Input Form".
  * legacy "Sheet1"/xls  (7 xlsx + 2 xls) -- older raw "Data Input Page" export.

We keep the SITE-SELECTION FACTORS and their per-factor Site Score, plus the
demographic components, target scores, labor, wash packages, traffic and
car-wash volume estimates.

PDFs are ignored (task: "the excels only").
"""
import os
import re
import csv
import glob

import openpyxl
import xlrd

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "old-proforma-data")
OUT_CSV = os.path.join(os.path.dirname(__file__), "..", "old-proforma-combined.csv")

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def norm(s):
    """normalise a label for fuzzy matching: uppercase, collapse whitespace/punct."""
    if s is None:
        return ""
    s = str(s).upper()
    s = re.sub(r"[^A-Z0-9%$]+", " ", s)
    return s.strip()

def num(v):
    """coerce to float if it looks numeric, else return the value/None unchanged."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("$", "").strip()
    try:
        return float(s)
    except ValueError:
        return None

def last_val(row):
    """last non-empty cell (read-only rows trail Nones we must skip)."""
    for v in reversed(row):
        if v is not None and v != "":
            return v
    return None


def parse_salesforce_id(fname):
    m = re.search(r"_(069[A-Za-z0-9]{12,15})\.(xlsx|xls|pdf)$", fname)
    return m.group(1) if m else ""

def parse_type(fname):
    low = fname.lower()
    if "flexserve" in low:
        return "Flexserve"
    if "flex proforma" in low or "flex_proforma" in low:
        return "Flex"
    if "express exterior" in low:
        return "Express Exterior"
    if "express" in low:
        return "Express"
    return "Other"


# The 10 site-selection factors, keyed by a normalised substring of the label
# found in column A of the Input Form.  order == appearance order in the sheet.
SITE_FACTORS = [
    ("area_profile",        "AREA PROFILE"),
    ("nearest_competition", "NEAREST COMPETITION"),
    ("weekly_hours",        "WEEKLY HOURS OF OPERATION"),
    ("type_of_site",        "TYPE OF SITE"),
    ("site_accessibility",  "SITE ACCESSIBILITY"),
    ("entrance_stack_up",   "ENTRANCE STACK UP"),
    ("free_vacuum_slots",   "NUMBER OF FREE VACUUM"),
    ("pay_stations",        "NUMBER OF PAY STATIONS"),
    ("visibility",          "VISIBILITY"),
    ("traffic_speed",       "TRAFFIC SPEED"),
]

DEMOG = [
    ("avg_household_size",  "AVERAGE HOUSEHOLD SIZE"),
    ("pct_pop_25_65",       "POPULATION 25 TO 65"),
    ("pct_hh_income_35k",   "HOUSEHOLDS WITH INCOMES"),
    ("base_price_carwash",  "BASE PRICE OF CAR WASH"),
]


# ---------------------------------------------------------------------------
# Input Form layout parser
# ---------------------------------------------------------------------------

def read_sheet_rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def parse_input_form(rows):
    """rows: list of row-lists from the 'Input Form' sheet."""
    rec = {}

    def find(pred, start=0):
        for i in range(start, len(rows)):
            if pred(rows[i]):
                return i
        return None

    def cell(r, c):
        row = rows[r] if 0 <= r < len(rows) else []
        return row[c] if c < len(row) else None

    def colA(row):
        return norm(row[0]) if row else ""

    def last_val(row):
        for v in reversed(row):
            if v is not None and v != "":
                return v
        return None

    # --- customer block -----------------------------------------------------
    i = find(lambda r: colA(r) == "CUSTOMER NAME")
    if i is not None:
        rec["customer_name"] = cell(i, 1)
    i = find(lambda r: colA(r) == "COMPANY NAME")
    if i is not None:
        rec["company_name"] = cell(i, 1)
    addr = []
    for lab in ("ADDRESS LINE 1", "ADDRESS LINE 2", "ADDRESS LINE 3"):
        i = find(lambda r, lab=lab: colA(r) == lab)
        if i is not None and cell(i, 1):
            addr.append(str(cell(i, 1)).strip())
    rec["address"] = ", ".join(addr)

    # weekly hours / avg daily wash hours (customer-info variant: col1 header)
    i = find(lambda r: colA(r) == "WEEKLY HOURS OF OPERATION"
             and norm(r[1] if len(r) > 1 else "") == "AVERAGE DAILY WASH HOURS")
    if i is not None:
        rec["weekly_hours_operation"] = num(cell(i + 1, 0))
        rec["avg_daily_wash_hours"] = num(cell(i + 1, 1))

    # --- labor --------------------------------------------------------------
    i = find(lambda r: colA(r) == "POSITION"
             and norm(r[1] if len(r) > 1 else "") == "LABOR HOURS")
    if i is not None:
        for off, key in ((1, "manager"), (2, "asst_manager"), (3, "attendants")):
            rec[f"labor_{key}_hours"] = num(cell(i + off, 1))
            rec[f"labor_{key}_wage"] = num(cell(i + off, 2))
            rec[f"labor_{key}_burden"] = num(cell(i + off, 3))

    # --- wash packages ------------------------------------------------------
    i = find(lambda r: colA(r) == "PACKAGES"
             and norm(r[1] if len(r) > 1 else "") == "PRICE")
    if i is not None:
        pkgs = [("basic", "BASIC PACKAGE"), ("menu1", "MENU PACKAGE 1"),
                ("menu2", "MENU PACKAGE 2"), ("menu3", "MENU PACKAGE 3"),
                ("menu4", "MENU PACKAGE 4")]
        # rows immediately following, matched by label
        for key, lab in pkgs:
            j = find(lambda r, lab=lab: colA(r) == lab, i)
            if j is not None:
                rec[f"pkg_{key}_price"] = num(cell(j, 1))
                rec[f"pkg_{key}_chem_cost"] = num(cell(j, 2))
                rec[f"pkg_{key}_pct_customers"] = num(cell(j, 3))

    # --- site-selection factors --------------------------------------------
    for key, lab in SITE_FACTORS:
        # header row whose colA matches and whose last col is "SITE SCORE"
        def is_hdr(r, lab=lab):
            if colA(r) != lab and not colA(r).startswith(lab):
                return False
            return any(norm(c) == "SITE SCORE" for c in r)
        h = find(is_hdr)
        if h is None:
            continue
        hdr = rows[h]
        # option columns are indices 1..4; site score is the col that == 'SITE SCORE'
        score_col = next((c for c, v in enumerate(hdr) if norm(v) == "SITE SCORE"), 5)
        sel = rows[h + 1] if h + 1 < len(rows) else []
        chosen = []
        for c in range(1, score_col):
            val = sel[c] if c < len(sel) else None
            if num(val) == 1.0:
                chosen.append(str(hdr[c]).strip() if c < len(hdr) and hdr[c] else f"col{c}")
        rec[f"factor_{key}_choice"] = " / ".join(chosen)
        rec[f"factor_{key}_score"] = num(sel[score_col] if score_col < len(sel) else None)

    i = find(lambda r: colA(r).startswith("CUMULATIVE SITE SCORE"))
    if i is not None:
        rec["cumulative_site_score"] = num(last_val(rows[i]))

    # --- demographic components --------------------------------------------
    for key, lab in DEMOG:
        i = find(lambda r, lab=lab: lab in colA(r))
        if i is not None:
            row = rows[i]
            # value is the last non-empty before score; layout: label ... value score
            rec[f"demog_{key}_value"] = num(cell(i, 4))
            rec[f"demog_{key}_score"] = num(cell(i, 5))
    i = find(lambda r: "CUMULATIVE DEMOGRAPHIC SCORE" in colA(r))
    if i is not None:
        rec["cumulative_demographic_score"] = num(cell(i, 5))

    # --- target scores ------------------------------------------------------
    for key, lab in (("first_year_target_score", "FIRST YEAR"),
                     ("second_year_target_score", "SECOND YEAR TARGET"),
                     ("mature_target_score", "MATURE TARGET")):
        i = find(lambda r, lab=lab: colA(r).startswith(lab))
        if i is not None:
            rec[key] = num(cell(i, 5))

    # --- traffic ------------------------------------------------------------
    i = find(lambda r: colA(r) == "TRAFFIC COUNT"
             and norm(r[1] if len(r) > 1 else "").startswith("YEAR 3"))
    if i is not None:
        rec["traffic_count"] = num(cell(i + 1, 0))
        rec["traffic_incr_y3"] = num(cell(i + 1, 1))
        rec["traffic_incr_y4"] = num(cell(i + 1, 2))
        rec["traffic_incr_y5"] = num(cell(i + 1, 3))

    # --- car wash volume estimates -----------------------------------------
    i = find(lambda r: colA(r) == "YEARLY")
    if i is not None:
        for y in range(1, 6):
            rec[f"vol_yearly_y{y}"] = num(cell(i, y))
    i = find(lambda r: colA(r) == "MONTHLY")
    if i is not None:
        for y in range(1, 6):
            rec[f"vol_monthly_y{y}"] = num(cell(i, y))

    # --- acquisition budget totals -----------------------------------------
    i = find(lambda r: colA(r) == "PROJECT COST")
    if i is not None:
        rec["project_cost"] = num(cell(i, 1))
    i = find(lambda r: colA(r) == "TOTAL DEBT")
    if i is not None:
        rec["total_debt"] = num(cell(i, 1))
    i = find(lambda r: colA(r) == "TOTAL LEASE PAYMENTS")
    if i is not None:
        rec["total_lease_payments"] = num(last_val(rows[i]))

    return rec


# ---------------------------------------------------------------------------
# legacy "Sheet1" / xls layout parser
#   label lives in column 1 (col 0 carries stray diagonal noise we ignore),
#   the 4 factor options sit in columns 2..5, the Site Score is the row's last
#   non-empty value, and factor labels/options are split across 2-3 rows.
#   The 10 factors appear in the SAME order as the Input Form.
# ---------------------------------------------------------------------------

def _first_after(row, idx, numeric):
    for c in range(idx + 1, len(row)):
        v = row[c]
        if numeric:
            n = num(v)
            if n is not None:
                return n
        else:
            if v not in (None, ""):
                return v
    return None


def _row_text(rows, r):
    return " | ".join("" if v in (None, "") else str(v) for v in rows[r])


def parse_legacy(rows):
    rec = {}

    def find_label(*keys):
        for r, row in enumerate(rows):
            for c, v in enumerate(row):
                nv = norm(v)
                if nv and all(k in nv for k in keys):
                    return r, c
        return None, None

    def scalar_after(numeric=True, *keys):
        r, c = find_label(*keys)
        if r is None:
            return None
        return _first_after(rows[r], c, numeric)

    # customer / address (repeated on every page; first hit is fine)
    rec["customer_name"] = scalar_after(False, "CUSTOMER")
    rec["address"] = scalar_after(False, "ADDRESS")

    # --- site-selection factors, matched positionally in appearance order ----
    start, _ = find_label("SITE", "SPECIFIC", "FACTORS")
    end, _ = find_label("CUMULATIVE", "SITE", "SCORE")
    found = []
    if start is not None and end is not None:
        buf = []
        for r in range(start + 1, end):
            row = rows[r][:8]
            has_marker = any(norm(v) in ("SITE", "SCORE") for v in row)
            sel_cols = [c for c in range(2, 6) if c < len(row) and num(row[c]) == 1.0]
            if sel_cols and not has_marker:
                score = last_val(row)
                labels = []
                for c in sel_cols:
                    parts = [str(b[c]).strip() for b in buf
                             if c < len(b) and b[c] not in (None, "")
                             and norm(b[c]) not in ("SITE", "SCORE")]
                    labels.append(" ".join(parts).strip())
                found.append((" / ".join([x for x in labels if x]), num(score)))
                buf = []
            else:
                buf.append(row)
    for idx, (key, _) in enumerate(SITE_FACTORS):
        if idx < len(found):
            rec[f"factor_{key}_choice"], rec[f"factor_{key}_score"] = found[idx]

    rec["cumulative_site_score"] = scalar_after(True, "CUMULATIVE", "SITE", "SCORE")

    # --- demographics: value = 1st numeric after label, score = 2nd ----------
    for key, keys in (("avg_household_size", ("AVERAGE", "HOUSEHOLD", "SIZE")),
                      ("pct_pop_25_65", ("POPULATION", "25", "65")),
                      ("pct_hh_income_35k", ("HOUSEHOLDS", "INCOMES")),
                      ("base_price_carwash", ("BASE", "PRICE", "CAR", "WASH"))):
        r, c = find_label(*keys)
        if r is not None:
            nums = [num(v) for v in rows[r][c + 1:] if num(v) is not None]
            if nums:
                rec[f"demog_{key}_value"] = nums[0]
            if len(nums) > 1:
                rec[f"demog_{key}_score"] = nums[1]

    rec["cumulative_demographic_score"] = scalar_after(True, "CUMULATIVE", "DEMOGRAPHIC")
    rec["first_year_target_score"] = scalar_after(True, "FIRST", "YEAR", "TARGET")
    rec["second_year_target_score"] = scalar_after(True, "SECOND", "YEAR", "TARGET")
    rec["mature_target_score"] = scalar_after(True, "MATURE", "TARGET")

    # traffic + volume
    rec["traffic_count"] = scalar_after(True, "TRAFFIC", "COUNT", "BOTH", "DIRECTIONS")
    for label, pref in (("YEARLY", "vol_yearly"), ("MONTHLY", "vol_monthly")):
        r, c = find_label(label)
        if r is not None:
            vals = [num(v) for v in rows[r][c + 1:]]
            vals = [v for v in vals if v is not None][:5]
            for y, v in enumerate(vals, 1):
                rec[f"{pref}_y{y}"] = v
    return rec


# ---------------------------------------------------------------------------
# driver
# ---------------------------------------------------------------------------

def process_file(path):
    fname = os.path.basename(path)
    rec = {
        "source_file": fname,
        "salesforce_id": parse_salesforce_id(fname),
        "proforma_type": parse_type(fname),
    }
    ext = fname.lower().rsplit(".", 1)[-1]
    if ext == "xlsx":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if "Input Form" in wb.sheetnames:
            rec["layout"] = "input_form"
            rec.update(parse_input_form(read_sheet_rows(wb["Input Form"])))
        else:
            rec["layout"] = "legacy_sheet1"
            rec.update(parse_legacy(read_sheet_rows(wb[wb.sheetnames[0]])))
        wb.close()
    elif ext == "xls":
        rec["layout"] = "legacy_xls"
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        rec.update(parse_legacy(rows))
    return rec


def main():
    paths = sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsx")) +
                   glob.glob(os.path.join(DATA_DIR, "*.xls")))
    records = []
    for p in paths:
        try:
            records.append(process_file(p))
        except Exception as e:
            print("ERROR", os.path.basename(p), e)
            records.append({"source_file": os.path.basename(p),
                            "salesforce_id": parse_salesforce_id(os.path.basename(p)),
                            "layout": "ERROR", "error": str(e)})

    # stable, grouped column order
    lead = ["source_file", "salesforce_id", "proforma_type", "layout",
            "customer_name", "company_name", "address",
            "weekly_hours_operation", "avg_daily_wash_hours"]
    labor = [f"labor_{k}_{m}" for k in ("manager", "asst_manager", "attendants")
             for m in ("hours", "wage", "burden")]
    pkg = [f"pkg_{k}_{m}" for k in ("basic", "menu1", "menu2", "menu3", "menu4")
           for m in ("price", "chem_cost", "pct_customers")]
    factors = []
    for key, _ in SITE_FACTORS:
        factors += [f"factor_{key}_choice", f"factor_{key}_score"]
    factors += ["cumulative_site_score"]
    demog = []
    for key, _ in DEMOG:
        demog += [f"demog_{key}_value", f"demog_{key}_score"]
    demog += ["cumulative_demographic_score",
              "first_year_target_score", "second_year_target_score", "mature_target_score"]
    traffic = ["traffic_count", "traffic_incr_y3", "traffic_incr_y4", "traffic_incr_y5"]
    vol = [f"vol_yearly_y{y}" for y in range(1, 6)] + [f"vol_monthly_y{y}" for y in range(1, 6)]
    acq = ["project_cost", "total_debt", "total_lease_payments"]
    cols = lead + labor + pkg + factors + demog + traffic + vol + acq
    # include any stragglers (e.g. error)
    extra = [c for r in records for c in r if c not in cols]
    for c in dict.fromkeys(extra):
        cols.append(c)

    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in records:
            w.writerow(r)

    print(f"Wrote {len(records)} rows -> {OUT_CSV}")
    by = {}
    for r in records:
        by[r.get("layout")] = by.get(r.get("layout"), 0) + 1
    print("by layout:", by)


if __name__ == "__main__":
    main()
