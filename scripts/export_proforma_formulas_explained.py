#!/usr/bin/env python3
"""Generate human-readable formula documentation for the Sonny's Proforma xlsx."""

from __future__ import annotations

import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def cell_text(ws, r: int, c: int) -> str:
    v = ws.cell(r, c).value
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if isinstance(v, str) and s.startswith("="):
        return ""
    return s


def input_row_label(ws, r: int, formula_col: int) -> str:
    """Best-effort label for Input Form rows."""
    if r == 9 and formula_col == 2:
        return "Average Daily Wash Hours (col B) — derived from Weekly Hours (col A)"
    # Site score grid: factor name is usually col A on row above the 1/0 row
    if 35 <= r <= 53 and formula_col == 6:
        a_above = cell_text(ws, r - 1, 1)
        if a_above:
            return f"Site score — {a_above}"
        return "Site score (factor row)"
    # Default: column A then B
    for col in (1, 2):
        t = cell_text(ws, r, col)
        if t:
            return t
    return f"Row {r}"


def proforma_row_label(ws, r: int) -> str:
    b = ws.cell(r, 2).value
    if b is None:
        a = cell_text(ws, r, 1)
        return a or f"Row {r}"
    s = str(b).strip()
    if isinstance(b, str) and s.startswith("="):
        inner = s[1:].strip()
        if inner.isidentifier():
            return f"[Named label: {inner} → package name text from Input Form]"
        return f"[Formula label: {inner[:50]}]"
    return s


def nearest_col_headers(ws, r: int, c: int) -> tuple[str, str]:
    """Scan upward for last non-formula text in this column."""
    title, tr = "", None
    for rr in range(r - 1, max(0, r - 80), -1):
        if rr < 1:
            break
        v = ws.cell(rr, c).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if isinstance(v, str) and s.startswith("="):
            continue
        title = s
        tr = rr
        break
    return title, tr or 0


def md_esc(x: object) -> str:
    return str(x).replace("|", "·").replace("\n", " ")


def explain_formula(
    formula: str,
    sheet: str,
    row_lbl: str,
    col_h: str,
    cell_ref: str,
    defined_name_set: set[str],
) -> str:
    """Short plain-English explanation (ordered: specific rules first)."""
    f = formula.strip()

    # True Excel names only (avoids treating =J84 as a "name")
    m = re.fullmatch(r"=([^!'\"]+)", f)
    if m:
        token = m.group(1).strip()
        if token in defined_name_set:
            return (
                f"Uses the **defined name `{token}`** — Excel’s shortcut to an Input Form cell "
                f"(see glossary above for the exact address)."
            )

    if f in ("=A9/7",) or re.fullmatch(r"=\s*A9\s*/\s*7", f):
        return "Divides **weekly operating hours** by 7 → **average hours open per day** (same number shown in col B)."

    if f.startswith("='Input Form'!"):
        return (
            "Pulls a value **from Input Form** so the pro forma **mirrors your assumptions** (no re-typing)."
        )

    if "PMT(" in f:
        return (
            "Computes a **monthly loan payment** (`PMT`): annual rate ÷ 12, term in months, loan balance. "
            "`*-1` flips Excel’s sign so the payment shows as a **positive** cash outflow."
        )

    if f.startswith("=SUM(F57:F60)/4"):
        return "**Average** of the four demographic scores (household size, age mix, income mix, price)."

    if re.fullmatch(r"=F35\+F37\+F39\+F41\+F43\+F45\+F47\+F49\+F51\+F53", f):
        return "**Total site score** — adds the 10 category point contributions above."

    if cell_ref == "F64" and sheet == "Input Form":
        return (
            "**First-year startup target score:** `(site score × 0.7) × (1 + demo adjustment) ÷ 85` — feeds "
            "column **B** annual volume."
        )

    if cell_ref == "F65" and sheet == "Input Form":
        return "**Year-2-style target score:** `site score × (1 + demo adjustment) ÷ 92` — used for column **C** annual volume path."

    if cell_ref == "F66" and sheet == "Input Form":
        return "**Mature target score:** `site score × (1 + demo adjustment) ÷ 76` — used for column **F** annual volume path."

    if cell_ref == "D74" and sheet == "Input Form":
        return (
            "**Annual cars (column D):** blends **site + demographic score** with the **Year 3** price/ramp "
            "inputs (`B70` in the Year 2–5 grid) × tunnel throughput × **300** days."
        )

    if cell_ref == "E74" and sheet == "Input Form":
        return (
            "**Annual cars (column E):** same idea as **D**, but uses the **Year 4** uplift column **`C70`** "
            "and divisor **82** in the template."
        )

    if cell_ref == "F74" and sheet == "Input Form":
        return "**Annual cars (column F)** uses the **mature score (F66)** × ramp (`D70`) × throughput × **300** days."

    if re.match(r"^=[CDEF]74/12$", f):
        return "Takes the **column’s annual wash forecast** and divides by 12 → **cars per month** for that scenario."

    if "300" in f and "A70" in f:
        return (
            "Builds **annual car count** from score × base tunnel throughput (A70) × 300 operating days, "
            "with year-specific uplift via B70–D70 where applicable."
        )

    if "/12" in f and "B74" in f:
        return "Converts **annual washes → average per month** (÷ 12)."

    if "/4.5/7*3" in f or "/4.5/7" in f:
        return (
            "Estimates **max cars in a peak operating sub-window**: monthly volume spread over ~4.5 "
            "“peak” days per week, 7 days, scaled to a 3-hour rush (rough capacity check vs hourly tunnel)."
        )

    if "IFERROR" in f and "/$B9" in f:
        return (
            "**Peak hourly volume estimate:** spreads max daily cars across average daily wash hours (B9), "
            "with a 1.3 intensity factor. IFERROR shields blank hour assumptions."
        )

    if re.search(r"=1-C\d+", f):
        return "**Equity share** of this line: 1 minus the financed fraction."

    if re.search(r"=C\d+\*B\d+", f) and sheet == "Input Form":
        return "**$ financed** = project cost × loan % for that line (principal borrowed)."

    if re.search(r"=D\d+\*B\d+", f):
        return "**$ equity** = project cost × (1 − loan %) for that line."

    if f.startswith("=IFERROR((I$65*"):
        return (
            "**Break-even column revenue for this package:** (required monthly cars to break even) × "
            "(package price) × (customer mix %). Wrapped in IFERROR→0."
        )

    if re.match(r"^=J\$65\*F\d+\*G\d+$", f):
        return (
            "**Year-one column revenue:** (Input Form **monthly wash volume** B75) × price × mix % — "
            "same structure as column I but using actual projected volume."
        )

    if "I76" in f and "*$F75" in f:
        return "**Bank & card fees:** −(card-eligible revenue × ~3% × effective fee factor)."

    if re.search(r"\$\s*G\s*\d+\s*\*\s*I76", f) or re.search(r"\$\s*G\s*\d+\s*\*\s*J76", f):
        return "**Variable OpEx line:** `% of revenue` in column G × **total revenue** (I76/J76) for that column."

    if re.search(r"IFERROR\s*\(\s*\(\s*[IJ]\d+\s*/\s*[IJ]74", f):
        return (
            "**Expense as % of wash revenue:** this line’s $ amount ÷ **total on-line revenue** (I74/J74); "
            "IFERROR→0 when revenue is zero."
        )

    if "INPUT FORM" in f.upper() and "B13" in f and "52/12" in f:
        return (
            "**Estimated monthly Labor $** from Input Form staffing: hourly × hours × burden, annualized "
            "pieces for managers vs attendants — rolled from labor grid (rows 13–15)."
        )

    if "IFERROR" in f and "F94" in f and "E100" in f and "J77" in f:
        return (
            "**Break-even monthly car count:** combines fixed debt/lease-like cash needs in the numerator "
            "(loan payments + leases + fixed pieces) and divides by **net $/car** implied by Year One revenue "
            "per car (column J). IFERROR→0 protects blanks."
        )

    if f.startswith("=IF(E57") or ("E57" in f and "2.101" in f):
        return "Scores **average household size** vs benchmark ~2.1; mild penalty if households are very large."

    if f.startswith("=IF(B") and ",IF(" in f and sheet == "Input Form":
        return (
            "**Site category score:** one of columns **B–E** should be `1` (your chosen scenario); "
            "nested `IF`s map that choice to a **point weight** for this factor."
        )

    if "0.55" in f and "E58" in f and f.startswith("=-"):
        return "Scores **% population age 25–65** vs a **55%** target (positive if above target)."

    if re.match(r"^=-0\.75\*\(0\.5-E59\)", f):
        return "Scores **% households over $35k income** vs a **50%** target."

    if f.startswith("=IF(") and "E60" in f and "10" in f:
        return "Scores **local base wash price** vs a **$10** reference (helpful if below 10, penalty if far above)."

    if f.startswith("=IFERROR(") and "I76" in f and "/I74" in f:
        return "**Percent of revenue** for this line: line amount ÷ total on-line revenue; IFERROR→0."

    mi, mj = re.fullmatch(r"=I(\d+)", f), re.fullmatch(r"=J(\d+)", f)
    if sheet == "Pro Forma" and mi and cell_ref.startswith("J"):
        return "**Same $ as column I, this row** — J is wired to **mirror** I so both scenarios stay aligned."
    if sheet == "Pro Forma" and mj and cell_ref.startswith("I"):
        return "**Same $ as column J, this row** — I is wired to **mirror** J (used where Labor$ is computed in J)."

    if "SUM(" in f:
        return "Adds the referenced cells (**subtotal / rollup**)."

    if re.search(r"I76\s*/\s*I65", f):
        return (
            "**Revenue per car — Break Even column:** **Total revenue** ÷ **break-even monthly cars** (I65). "
            "IFERROR→0 if I65 is blank."
        )

    if re.search(r"J76\s*/\s*\(\s*J65", f) or "J76/(J65" in f.replace(" ", ""):
        return (
            "**Revenue per car — Year One column:** **Total revenue** ÷ **Year One monthly washes** (~`J65`, from Input **B75**). "
            "Adds tiny `0.000000001` to avoid divide-by-zero."
        )

    if re.search(r"-\s*\(\s*I74\s*\*\s*\$F75", f) or re.search(r"-\s*\(\s*J74\s*\*\s*\$F75", f):
        return (
            "**Merchant / bank fees:** treats a slice of **revenue** as subject to the fee inputs in **F75** and **G75** "
            "(negative sign = cash cost)."
        )

    if re.match(r"^=\s*I\$65\*G\d+\*F\d+$", f.replace(" ", "")):
        return (
            "**Package revenue — Break Even (no IFERROR on this row):** **I65** cars × **price** (F) × **mix %** (G)."
        )

    if "B75" in f and "Input Form" in f and "C19" in f and len(f) > 120:
        return (
            "**Chemical $ (Year One):** rebuilds `monthly volume × package mix × $/car chemical cost` for **each menu item** "
            "from Input Form and sums them — matches **wash mix** to **chemical cost per car**."
        )

    if "SUM(" in f and "Input Form" in f and "B75" in f and "D19" in f:
        return (
            "**Chemical $ (alternate build):** expands `B75 × price × mix × chemical $` across packages — same intent as the row above."
        )

    if "I76-I92" in f or "J76-J92" in f:
        return "**Net operating cash flow** = revenue after card fees minus operating expenses."

    if col_h.strip() == "Break Even":
        return "Amount for the **Break Even** scenario column (uses **break-even wash count** logic in I65 vs Year 1 volume)."

    if col_h.strip() == "Year One":
        return "Same line for the **Year One** scenario column — uses **Input Form monthly volume** (`B75`) where applicable."

    if col_h.strip() == "Percentage of Revenue":
        return "Shows this row as a **% of revenue** for the paired dollar column to the left."

    return (
        "**What to do in Excel:** trace colored precedents from this cell, or use **Evaluate Formula**. "
        "This line rolls up other cells—see the formula text for exact references."
    )


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    xlsx = root / (
        "Copy of Sonny's Express Exterior Proforma - 3301 W Hallandale Beach Blvd "
        "Pembroke Park, FL 33023.xlsx"
    )
    out = root / "Copy_of_Sonnys_Proforma_formulas_explained.md"
    wb = load_workbook(xlsx, data_only=False)
    dns = set(wb.defined_names.keys())

    lines: list[str] = []
    lines += [
        "# Sonny’s Express Exterior Proforma — formulas explained",
        "",
        "This document lists **every cell formula** with: **where it lives**, **what it’s for (name/label)**, "
        "and a **plain-English reading** plus a **tiny example** where it helps.",
        "",
        "---",
        "",
        "## Big picture (read this first)",
        "",
        "1. **Input Form** — All assumptions: customer info, **weekly hours → daily hours**, **site score** "
        "(pick-one buttons), **demographic tweaks**, **target scores**, **annual/monthly/daily/hourly wash "
        "volume** under several year profiles (columns B–F), **construction/debt split**, **loan payments**, "
        "and **lease adders**.",
        "",
        "2. **Pro Forma** — A presentation / reporting sheet. It **links** to Input Form, repeats **customer "
        "address blocks**, and builds **monthly P&L–style sections**.",
        "",
        "3. **Columns I vs J (most financial rows)** — Usually **“Break Even” vs “Year One”** under a "
        "“Monthly Wash Volume” style header (see **row 64** and again **row 68**). "
        "Lower in the sheet, column **I** may show dollar flows while **J** shows **“Percentage of Revenue”** "
        "(see **row 165**). Always read the **nearest header row above** your cell.",
        "",
        "4. **Named ranges** — Cells like `=CustomerName` are **Excel names** pointing to Input Form cells "
        "(table in the next section).",
        "",
        "---",
        "",
        "## Glossary: workbook names → what they mean",
        "",
        "| Excel name | Points to (typical) |",
        "|------------|---------------------|",
    ]
    for name, dfn in sorted(wb.defined_names.items(), key=lambda x: x[0].lower()):
        txt = getattr(dfn, "attr_text", "") or ""
        if txt.startswith("#"):
            lines.append(f"| `{name}` | *(broken ref in file: {txt})* |")
        else:
            lines.append(f"| `{name}` | `{txt}` |")
    lines += ["", "---", ""]

    # --- Input Form ---
    ws_in = wb["Input Form"]
    lines += ["## Sheet: **Input Form**", ""]
    input_sections = [
        (1, 10, "### Customer & operating hours"),
        (33, 54, "### Site-specific factor scores (rows 33–54)"),
        (56, 66, "### Demographic adjustments & target scores"),
        (70, 78, "### Volume ladder: annual → monthly → daily → hourly"),
        (79, 102, "### Project cost, financing & leases"),
    ]
    for lo, hi, title in input_sections:
        lines.append(title)
        lines.append("")
        lines.append("| Cell | Applies to | Formula | What it’s doing | Example |")
        lines.append("|------|------------|---------|-----------------|--------|")
        for r in range(lo, hi + 1):
            for c in range(1, ws_in.max_column + 1):
                v = ws_in.cell(r, c).value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                ref = f"{get_column_letter(c)}{r}"
                lbl = input_row_label(ws_in, r, c)
                ex = ""
                if ref == "B9":
                    ex = "If **A9** weekly hours = 84 → **B9** = 12 hours/day average."
                elif ref == "F54":
                    ex = "Adds 10 factor scores into one site total."
                elif "PMT" in v:
                    ex = "Like a mortgage payment on each financed slice."
                elif ref.startswith("F") and r in range(35, 54, 2):
                    ex = "Only one of B–E should be 1; the formula returns that row’s point weight."
                elif ref in ("B74", "C74"):
                    ex = "Cars/year = score × throughput × days."
                lines.append(
                    "| `{ref}` | {lbl} | `{v}` | {exp} | {ex} |".format(
                        ref=ref,
                        lbl=md_esc(lbl),
                        v=md_esc(v),
                        exp=md_esc(explain_formula(v, "Input Form", lbl, "", ref, dns)),
                        ex=md_esc(ex),
                    )
                )
        lines.append("")

    # --- Pro Forma ---
    ws_pf = wb["Pro Forma"]
    lines += ["## Sheet: **Pro Forma**", ""]
    lines += [
        "The Pro Forma repeats **the same financial pattern** many times for different **annual wash "
        "volume columns** (pulled from Input Form `B74`, `C74`, …) and for **break-even vs year-one** "
        "views. If two cells show the *same formula shape*, the *business meaning* is the same — only "
        "the scenario column or volume anchor changes.",
        "",
        "| Cell | Row label (col B) | Column meaning | Formula | Plain English |",
        "|------|-------------------|----------------|---------|---------------|",
    ]

    for row in ws_pf.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            r, c = cell.row, cell.column
            ref = f"{get_column_letter(c)}{r}"
            rl = proforma_row_label(ws_pf, r)
            ch, _ = nearest_col_headers(ws_pf, r, c)
            expl = explain_formula(v, "Pro Forma", rl, ch, ref, dns)
            # Widen column meaning for I/J
            col_meaning = ch or f"Column {get_column_letter(c)}"
            if c == 9 and not ch:
                col_meaning = "Usually **Break Even** or **Year One $** (see header above)"
            if c == 10 and not ch:
                col_meaning = "Usually **Year One $** or **% of revenue** (see header above)"
            lines.append(
                "| `{ref}` | {rl} | {cm} | `{v}` | {e} |".format(
                    ref=ref,
                    rl=md_esc(rl),
                    cm=md_esc(col_meaning),
                    v=md_esc(v),
                    e=md_esc(expl),
                )
            )

    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out} ({len(lines)} lines)")


if __name__ == "__main__":
    main()
